import uuid
from uuid import UUID
from datetime import datetime
from django.conf import settings
from django.utils import timezone
from urllib.parse import urlencode
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse, reverse_lazy
from django.http.response import HttpResponseRedirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import CreateView, UpdateView, DeleteView, DetailView, ListView, TemplateView

import openpyxl
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as ExcelImage

from publicmanager.saude_cadastro.models import CID
from publicmanager.saude_atendimento.models import Paciente, AnamnesePaciente, BoletimPaciente, ClassificacaoRisco, AtendimentoMedico
from publicmanager.saude_atendimento.forms import (
    PacienteCriarForm, PacienteAtualizarForm, PacienteCriarForm, PacienteAtualizarForm,
    AnamnesePacienteForm, PrincipioAtivo
)
from publicmanager.comum.utils import CheckUserTypeMixin
from publicmanager.saude_atendimento.forms import RelatorioAtendimentosForm

class AdmissaoPacienteListView(CheckUserTypeMixin, LoginRequiredMixin, ListView):
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    model = Paciente
    context_object_name = 'pacientes'
    template_name = 'saude_atendimento/urgencia/admissao_paciente_list.html'
    paginate_by = 10

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo') == 'urgencia' or self.request.GET.get('modulo') == 'consultorio':
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def get_queryset(self):
        queryset = super().get_queryset().filter(situacao='ATIVO').order_by('nome_paciente')

        if self.request.GET.get('buscar_nome') and self.request.GET.get('opcao_filtro') == "paciente":
            queryset = queryset.filter(nome_paciente__icontains=self.request.GET.get('buscar_nome'))
            
        elif self.request.GET.get('buscar_nome') and self.request.GET.get('opcao_filtro') == "cpf":
            queryset = queryset.filter(cpf=self.request.GET.get('buscar_nome'))

        elif self.request.GET.get('buscar_nome') and self.request.GET.get('opcao_filtro') == "rg":
            queryset = queryset.filter(rg=self.request.GET.get('buscar_nome'))

        elif self.request.GET.get('buscar_nome') and self.request.GET.get('opcao_filtro') == "sus":
            queryset = queryset.filter(cartao_sus=self.request.GET.get('buscar_nome'))
      
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        base_url = reverse('saude_atendimento:admissao_paciente_list')
        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})

        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})
        
        url = f'{base_url}?{query_string}'

        context['crumbs'] = [
            {'label': 'Listagem de Pacientes', 'url': url},
        ]

        context['modulo'] = self.request.GET.get('modulo', "")
        context['buscar_nome'] = self.request.GET.get('buscar_nome', "")
        context['opcao_filtro'] = self.request.GET.get('opcao_filtro', "") if self.request.GET.get('opcao_filtro', "") else "paciente"

        return context

class AdmissaoPacienteCreateView( CheckUserTypeMixin, LoginRequiredMixin, SuccessMessageMixin, CreateView):
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    model = Paciente
    context_object_name = "paciente"
    template_name = 'saude_atendimento/urgencia/admissao_paciente_create_update.html'
    form_class = PacienteCriarForm
    success_url = reverse_lazy('saude_atendimento:paciente_list')

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo') == 'urgencia' or self.request.GET.get('modulo') == 'consultorio':
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        list_url = reverse('saude_atendimento:admissao_paciente_list')
        add_url = reverse('saude_atendimento:admissao_paciente_add')

        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})

        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})

        context["view"] = "paciente_create_view"
        context['crumbs'] = [
            {'label': 'Listagem de Pacientes', 'url': f'{list_url}?{query_string}'},
            {'label': 'Cadastro de Paciente', 'url': f'{add_url}?{query_string}'},
        ]
        context['modulo'] = self.request.GET.get('modulo', "")

        return context

    def form_valid(self, form):
        anamnese_paciente = AnamnesePaciente()
        anamnese_paciente.save()

        paciente = form.save(commit=False)
        paciente.anamnese_paciente_id = anamnese_paciente.id

        if paciente.cpf:
            paciente.cpf = paciente.cpf.replace('.', '').replace('-', '')
        
        if paciente.rg:
            paciente.rg = paciente.rg.replace('.', '').replace('-', '')
        
        paciente.save()

        self.object = form.save(commit=False)
        self.object.unidade_saude=self.request.user.get_unidade_login().unidade
        self.object.save()

        base_url = reverse('saude_atendimento:admissao_paciente_update', kwargs={'pk': self.object.pk})
        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})
            url = f'{base_url}?{query_string}'
        
        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})
            url = f'{base_url}?{query_string}'
            
        messages.success(self.request, 'Pacientes criado com sucesso!', extra_tags=f"paciente_criado_com_sucesso" )
        
        return redirect(url)

class AdmissaoPacienteUpdateView(CheckUserTypeMixin, LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    model = Paciente
    context_object_name = "paciente"
    template_name = 'saude_atendimento/urgencia/admissao_paciente_create_update.html'
    form_class = PacienteAtualizarForm

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo') == 'urgencia' or self.request.GET.get('modulo') == 'consultorio':
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        list_url = reverse('saude_atendimento:admissao_paciente_list')
        update_url = reverse('saude_atendimento:admissao_paciente_update', kwargs={'pk': self.kwargs.get('pk')})

        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})

        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})

        context["view"] = "paciente_update_view"
        context['crumbs'] = [
            {'label': 'Listagem de Pacientes', 'url': f'{list_url}?{query_string}'},
            {'label': 'Edição de dados do Paciente', 'url': f'{update_url}?{query_string}'},
        ]

        context["id_anamnese_paciente"] = self.object.anamnese_paciente_id
        context['modulo'] = self.request.GET.get('modulo', "")

        return context
    
    def get_success_url(self):
        base_url = reverse('saude_atendimento:admissao_paciente_update', kwargs={'pk': self.object.pk})
        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})
            url = f'{base_url}?{query_string}'
        
        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})
            url = f'{base_url}?{query_string}'

        messages.success(self.request, 'Paciente atualizado com sucesso!', extra_tags=f"paciente_atualizado_com_sucesso" )
        
        return url
    
    def form_valid(self, form):
        paciente = form.save(commit=False)
        
        if paciente.cpf:
            paciente.cpf = paciente.cpf.replace('.', '').replace('-', '')
        
        if paciente.rg:
            paciente.rg = paciente.rg.replace('.', '').replace('-', '')
        
        paciente.save()

        self.object = form.save(commit=False)
        self.object.unidade_saude=self.request.user.get_unidade_login().unidade
        self.object.save()

        return HttpResponseRedirect(self.get_success_url())

class AdmissaoPacienteDeleteView(CheckUserTypeMixin, LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    model = Paciente
    template_name = "saude_atendimento/urgencia/admissao_paciente_list.html"
    success_message = 'Seus dados foram deletados conforme solicitado.'

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo') == 'urgencia' or self.request.GET.get('modulo') == 'consultorio':
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['modulo'] = self.request.GET.get('modulo', "")

        return context

    def get_success_url(self):
        base_url = reverse('saude_atendimento:admissao_paciente_list')
        modulo = self.request.GET.get('modulo')

        if modulo in ['urgencia', 'consultorio']:
            query_string = urlencode({'modulo': modulo})
            url = f'{base_url}?{query_string}'
        else:
            url = base_url

        return url

class RelatorioPacientePDFView(CheckUserTypeMixin, LoginRequiredMixin, TemplateView):
    template_name = "exportacao/paciente_pdf.html"
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]

    def get_context_data(self, **kwargs):
        context = super(RelatorioPacientePDFView, self).get_context_data(**kwargs)

        data_hora_atual = datetime.now()

        context['data_impressao'] = data_hora_atual.strftime("%d/%m/%d %H:%M")
        context['paciente'] = Paciente.objects.get(pk=kwargs.get('pk'))

        return context

class AnamnesePacienteCreateView(CheckUserTypeMixin, LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = AnamnesePaciente
    template_name = 'saude_atendimento/urgencia/anamnese_create_update_form.html'
    form_class = AnamnesePacienteForm
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA]
    success_message = 'Seus dados foram cadastrado conforme preenchimento do formulário.'

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo'):
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def get_success_url(self):
        base_url = reverse('saude_atendimento:admissao_paciente_list')
        modulo = self.request.GET.get('modulo')

        if modulo in ['urgencia', 'consultorio']:
            query_string = urlencode({'modulo': modulo})
            url = f'{base_url}?{query_string}'
        else:
            url = base_url

        return url

class AnamnesePacienteUpdateView(CheckUserTypeMixin, LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = AnamnesePaciente
    context_object_name = "anamnese_paciente"
    template_name = 'saude_atendimento/urgencia/anamnese_create_update_form.html'
    form_class = AnamnesePacienteForm
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA]
    # success_url = reverse_lazy('saude_atendimento:admissao_paciente_list')
    success_message = 'Seus dados foram atualizados conforme preencimento do formulário.'

    def dispatch(self, request, *args, **kwargs):
        response = super().dispatch(request, *args, **kwargs)
        if not request.user.is_authenticated or request.user.is_anonymous:
            return response

        if self.request.GET.get('modulo'):
            return super().dispatch(request, *args, **kwargs)
            
        messages.warning(self.request, 'Por favor, forneça uma URL válida.')
        return redirect(reverse('dashboard:index'))

    def form_valid(self, form):
        instance = form.save(commit=True)
       
        if self.request.POST.get('alergias_medicamentosas'):
            alergias_medicamentosas = [UUID(uuid) for uuid in self.request.POST.get('alergias_medicamentosas').split(',')]
            instance.alergias_medicamentosas.set(alergias_medicamentosas)
        else:
            instance.alergias_medicamentosas.set([])

        if self.request.POST.get('antecedentes_patologicos_pessoais'):
            antecedentes_patologicos_pessoais = [UUID(uuid) for uuid in self.request.POST.get('antecedentes_patologicos_pessoais').split(',')]
            instance.antecedentes_patologicos_pessoais.set(antecedentes_patologicos_pessoais)
        else:
            instance.antecedentes_patologicos_pessoais.set([])

        if self.request.POST.get('antecedentes_patologicos_familiares'):
            antecedentes_patologicos_familiares = [UUID(uuid) for uuid in self.request.POST.get('antecedentes_patologicos_familiares').split(',')]
            instance.antecedentes_patologicos_familiares.set(antecedentes_patologicos_familiares)
        else:
            instance.antecedentes_patologicos_familiares.set([])
        
        instance.save()
        
        return super().form_valid(form)
    

    def get_context_data(self, **kwargs):
        boletim_id = self.request.GET.get('boletim')

        context = super().get_context_data(**kwargs)

        query_string = urlencode({'modulo': self.request.GET.get('modulo', "")})

        list_pacientes_url = reverse('saude_atendimento:admissao_paciente_list')
        update_paciente_url = reverse('saude_atendimento:admissao_paciente_update', kwargs={'pk': self.object.paciente.pk})
        anamnese_url = reverse('saude_atendimento:anamnese_update', kwargs={'pk': self.object.pk})


        context["cids"] = CID.objects.all()
        context["principios_ativo"] = PrincipioAtivo.objects.all()
        context['modulo'] = self.request.GET.get('modulo', "")
        context['boletim'] = self.request.GET.get('boletim', "")
        context['paciente'] = self.request.GET.get('paciente', "")


        if ClassificacaoRisco.objects.filter(boletim=boletim_id).exists():
            classificacaorisco = ClassificacaoRisco.objects.get(boletim=boletim_id)
            base_url = reverse('saude_atendimento:classificacao_risco_update', kwargs={'pk': classificacaorisco.id})
        else:
            base_url = reverse('saude_atendimento:classificacao_risco_create', kwargs={'pk': boletim_id})    

        context['url_redirect'] = f'{base_url}?paciente={self.object.paciente.id}'

        context['crumbs'] = [
            {'label': 'Listagem de Pacientes', 'url': f'{list_pacientes_url}?{query_string}'},
            {'label': 'Editar dados do paciente', 'url': f'{update_paciente_url}?{query_string}'},
            {'label': 'Anamnese do Paciente', 'url': f'{anamnese_url}?{query_string}'}
        ]
        
        return context
    
    def get_success_url(self):
        boletim_id = self.request.GET.get('boletim')
        
        if ClassificacaoRisco.objects.filter(boletim=boletim_id).exists():
            classificacaorisco = ClassificacaoRisco.objects.get(boletim=boletim_id)
            base_url = reverse('saude_atendimento:classificacao_risco_update', kwargs={'pk': classificacaorisco.id})
        else:
            base_url = reverse('saude_atendimento:classificacao_risco_create', kwargs={'pk': boletim_id})    
        modulo = self.request.GET.get('modulo')

        if modulo in ['urgencia', 'consultorio']:
            query_string = urlencode({'modulo': modulo})
            url = f'{base_url}?{query_string}&paciente={self.object.paciente.id}'
        else:
            url = f'{base_url}?paciente={self.object.paciente.id}'

        return url

class PacienteHistoricoListView(CheckUserTypeMixin, LoginRequiredMixin, ListView):
    model = BoletimPaciente
    # required_permissions = [settings.DESENVOLVEDOR, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    template_name = "saude_atendimento/urgencia/paciente_historico_list.html"
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset().filter(
            unidade_saude=self.request.user.get_unidade_login().unidade,
            paciente=self.kwargs.get('pk'),
        ).exclude(situacao__in=[BoletimPaciente.EM_ABERTO, BoletimPaciente.EM_ANDAMENTO]).order_by('-created_at')
      
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['data_atual'] = timezone.now()
        context['slc_paciente'] = Paciente.objects.get(pk=self.kwargs.get('pk'))

        list_url = reverse('saude_atendimento:admissao_paciente_list')
        historico_url = reverse('saude_atendimento:paciente_historico_list', kwargs={'pk': self.kwargs.get('pk')})

        if self.request.GET.get('modulo') == 'urgencia':
            query_string = urlencode({'modulo': 'urgencia'})

        elif self.request.GET.get('modulo') == 'consultorio':
            query_string = urlencode({'modulo': 'consultorio'})

        context['crumbs'] = [
            {'label': 'Listagem de Pacientes', 'url': f'{list_url}?{query_string}'},
            {'label': 'Histórico de Atendimento', 'url': f'{historico_url}?{query_string}'},
        ]

        context['modulo'] = self.request.GET.get('modulo', "")
  
        return context

class PacienteHistoricoTimeLineListView(CheckUserTypeMixin, LoginRequiredMixin, DetailView):
    model = BoletimPaciente
    context_object_name = "boletim"
    # required_permissions = [settings.DESENVOLVEDOR, settings.MEDICO, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    template_name = "saude_atendimento/urgencia/paciente_historico_timeline.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['data_atual'] = timezone.now()

        return context

class PacienteHistoricoFichaListView(CheckUserTypeMixin, LoginRequiredMixin, DetailView):
    model = BoletimPaciente
    context_object_name = "boletim"
    # required_permissions = [settings.DESENVOLVEDOR, settings.MEDICO, settings.RECEPCIONISTA, settings.ADMINISTRADO]
    template_name = "saude_atendimento/urgencia/paciente_historico_ficha.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['data_atual'] = timezone.now()
        context['title_relatorio'] = 'INFORMAÇÕES DO ATENDIMENTO MÉDICO'
        context['classificacao_risco'] = ClassificacaoRisco.objects.filter(boletim__id=self.object.id, status=ClassificacaoRisco.ATIVO).first()
        context['unidade_saude'] = self.request.user.get_unidade_login().unidade

        return context
    
class RastreamentoPacientesListView(CheckUserTypeMixin, LoginRequiredMixin, TemplateView):
    template_name = 'saude_atendimento/urgencia/rastreamento_pacientes_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['unidade_saude'] = self.request.user.get_unidade_login().unidade

        return context


def exportar_relatorio_atendimentos(request):
    if request.method == 'POST':
        form = RelatorioAtendimentosForm(request.POST)
        if form.is_valid():
            data_inicial = form.cleaned_data['data_inicial']
            data_final = form.cleaned_data['data_final']
            unidade_saude = form.cleaned_data['unidade_saude']

            unique_id = uuid.uuid4()
            filename = f"relatorio_atendimentos_{unique_id}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Atendimentos'

            # Adicione a logo
            logo_path = "publicmanager/comum/static/img/logo-1gov.png"
            logo = ExcelImage(logo_path)
            logo.height = 90
            logo.width = 150

            # Altura da linha e a largura da coluna para a logo
            ws.row_dimensions[1].height = 70
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.add_image(logo, 'A1')

            # Adicione o título
            title_font = Font(size=14, bold=True)
            ws.merge_cells('B1:D1')
            title_cell = ws['B1']
            title_cell.value = 'Relatório de Atendimentos'
            title_cell.font = title_font

            # Escreve o cabeçalho do XLSX
            columns = ['Data', 'Paciente', 'Unidade de Saúde', 'Profissional']
            header_font = Font(bold=True)

            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = column_title
                cell.font = header_font

            total = 0
            atendimentos = AtendimentoMedico.objects.none()
            if data_inicial and data_final and unidade_saude:
                atendimentos = AtendimentoMedico.objects.filter(unidade_saude=unidade_saude, created_at__date__range=(data_inicial, data_final)).order_by('created_at')

            if atendimentos:
                for row_num, atendimento in enumerate(atendimentos, 4):
                    profissionais = ', '.join([profissional.nome_profissional for profissional in atendimento.profissionais.all()]) if atendimento.profissionais.exists() else '-----'

                    ws.cell(row=row_num, column=1, value=atendimento.created_at.strftime('%d/%m/%Y %H:%M'))
                    ws.cell(row=row_num, column=2, value=atendimento.lista_chamada.paciente.nome_paciente)
                    ws.cell(row=row_num, column=3, value=atendimento.lista_chamada.unidade_saude.nome)
                    ws.cell(row=row_num, column=4, value=profissionais)
                    total += 1

                ws.cell(row=row_num + 1, column=4, value='Total')
                ws.cell(row=row_num + 2, column=4, value=total)
            
            messages.info(request, 'Relatório exportado com sucesso!')

            # Salva o arquivo em memória
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Cria a resposta HTTP para download
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    else:
        form = RelatorioAtendimentosForm()

    return render(request, 'admin/exportar_relatorio_atendimentos.html', {'form': form})